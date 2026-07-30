import tkinter as tk
import customtkinter as ctk
from tkinter import messagebox, filedialog
import openpyxl
from date_utils import to_display, DateEntry
from database import (get_vendors, get_vendor_outstanding_report,
                      get_purchases, get_notes, get_vouchers, get_expenses)
from searchable_combo import SearchableComboBox


def fmt(v):
    try:
        return f"\u20b9{float(v):,.2f}"
    except Exception:
        return str(v)


class ReportFrame(ctk.CTkFrame):
    def __init__(self, master, db_path, fy_id=None, **kwargs):
        super().__init__(master, fg_color="#F4F6FB")
        self.db_path = db_path
        self.fy_id = fy_id
        self._build()

    def _build(self):
        for w in self.winfo_children():
            w.destroy()

        ctk.CTkLabel(self, text="Reports",
                     font=ctk.CTkFont(size=22, weight="bold"),
                     text_color="#1B2B6B").pack(anchor="w", padx=24, pady=(20, 8))

        self._tab = ctk.CTkTabview(self, fg_color="white", corner_radius=12)
        self._tab.pack(fill="both", expand=True, padx=24, pady=(0, 24))
        self._tab.add("Transaction Reports")
        self._tab.add("Outstanding Report")
        self._tab.add("Invoice Settlement")

        self._build_transactions(self._tab.tab("Transaction Reports"))
        self._build_outstanding(self._tab.tab("Outstanding Report"))
        self._build_settlement(self._tab.tab("Invoice Settlement"))

    # ── Transaction Reports (Payment / Purchase / Transfer / Notes / Expense) ──

    def _payment_rows(self, df, dt, vid, cancelled):
        rows = get_vouchers(self.db_path, date_from=df, date_to=dt,
                            vendor_id=vid, vtype="payment",
                            include_cancelled=cancelled, fy_id=self.fy_id)
        return [[to_display(v["date"]) if v.get("date") else "-",
                 v["voucher_no"], v.get("vendor_name") or "-",
                 v.get("from_account_name") or "-", v.get("payment_mode") or "-",
                 v["amount"], v.get("narration") or "-", v.get("status", "active")]
                for v in rows]

    def _transfer_rows(self, df, dt, vid, cancelled):
        rows = get_vouchers(self.db_path, date_from=df, date_to=dt,
                            vtype="transfer", include_cancelled=cancelled,
                            fy_id=self.fy_id)
        return [[to_display(v["date"]) if v.get("date") else "-",
                 v["voucher_no"], v.get("from_account_name") or "-",
                 v.get("to_account_name") or "-", v["amount"],
                 v.get("narration") or "-", v.get("status", "active")]
                for v in rows]

    def _purchase_rows(self, df, dt, vid, cancelled):
        rows = get_purchases(self.db_path, vendor_id=vid, date_from=df, date_to=dt,
                             include_cancelled=cancelled, fy_id=self.fy_id)
        return [[to_display(p["date"]) if p.get("date") else "-",
                 p["voucher_no"], p.get("vendor_name") or "-",
                 p.get("invoice_number") or "-",
                 to_display(p["invoice_date"]) if p.get("invoice_date") else "-",
                 p["purchase_value"], p["gst_amount"], p["total_value"],
                 p["outstanding"], p.get("status", "active")]
                for p in rows]

    def _note_rows(self, table, df, dt, vid, cancelled):
        rows = get_notes(self.db_path, table, vendor_id=vid, date_from=df, date_to=dt,
                         include_cancelled=cancelled, fy_id=self.fy_id)
        return [[to_display(n["date"]) if n.get("date") else "-",
                 n["voucher_no"], n.get("vendor_name") or "-",
                 n["value"], n["gst_amount"], n["total_value"],
                 n.get("narration") or "-", n.get("status", "active")]
                for n in rows]

    def _expense_rows(self, df, dt, vid, cancelled):
        rows = get_expenses(self.db_path, date_from=df, date_to=dt,
                            include_cancelled=cancelled, fy_id=self.fy_id)
        return [[to_display(e["date"]) if e.get("date") else "-",
                 e["voucher_no"], e.get("expense_head_name") or "-",
                 e.get("account_name") or "-", e.get("payment_mode") or "-",
                 e["amount"], e.get("narration") or "-", e.get("status", "active")]
                for e in rows]

    def _txn_report_configs(self):
        return {
            "Payment Vouchers": {
                "headers": ["Date", "Voucher No", "Vendor / Payee", "From Account",
                            "Mode", "Amount", "Narration", "Status"],
                "widths":  [85, 95, 155, 140, 90, 110, 210, 75],
                "numeric": [5], "vendor_filter": True, "fetch": self._payment_rows,
            },
            "Internal Transfers": {
                "headers": ["Date", "Voucher No", "From Account", "To Account",
                            "Amount", "Narration", "Status"],
                "widths":  [85, 95, 155, 155, 110, 240, 75],
                "numeric": [4], "vendor_filter": False, "fetch": self._transfer_rows,
            },
            "Purchases": {
                "headers": ["Date", "Voucher No", "Vendor", "Invoice No", "Invoice Date",
                            "Purchase Value", "GST", "Total Value", "Outstanding", "Status"],
                "widths":  [80, 90, 140, 105, 95, 110, 90, 110, 100, 75],
                "numeric": [5, 6, 7, 8], "vendor_filter": True, "fetch": self._purchase_rows,
            },
            "Credit Notes": {
                "headers": ["Date", "Voucher No", "Vendor", "Value", "GST",
                            "Total Value", "Narration", "Status"],
                "widths":  [85, 95, 150, 100, 90, 110, 210, 75],
                "numeric": [3, 4, 5], "vendor_filter": True,
                "fetch": lambda df, dt, vid, c: self._note_rows("credit_notes", df, dt, vid, c),
            },
            "Debit Notes": {
                "headers": ["Date", "Voucher No", "Vendor", "Value", "GST",
                            "Total Value", "Narration", "Status"],
                "widths":  [85, 95, 150, 100, 90, 110, 210, 75],
                "numeric": [3, 4, 5], "vendor_filter": True,
                "fetch": lambda df, dt, vid, c: self._note_rows("debit_notes", df, dt, vid, c),
            },
            "Expenses": {
                "headers": ["Date", "Voucher No", "Expense Head", "Account",
                            "Mode", "Amount", "Narration", "Status"],
                "widths":  [85, 95, 150, 140, 90, 110, 200, 75],
                "numeric": [5], "vendor_filter": False, "fetch": self._expense_rows,
            },
        }

    def _build_transactions(self, parent):
        self._txn_configs = self._txn_report_configs()
        self._txn_rows = []

        top = ctk.CTkFrame(parent, fg_color="transparent")
        top.pack(fill="x", pady=(10, 4))

        ctk.CTkLabel(top, text="Report:").pack(side="left", padx=(8, 4))
        self._txn_type_var = tk.StringVar(value="Payment Vouchers")
        ctk.CTkOptionMenu(top, values=list(self._txn_configs.keys()),
                         variable=self._txn_type_var, width=160,
                         command=self._on_txn_type_change).pack(side="left", padx=4)

        ctk.CTkLabel(top, text="From").pack(side="left", padx=(14, 4))
        self._txn_df_entry = DateEntry(top, initial_date="")
        self._txn_df_entry.pack(side="left", padx=(0, 8))

        ctk.CTkLabel(top, text="To").pack(side="left", padx=(0, 4))
        self._txn_dt_entry = DateEntry(top, initial_date="")
        self._txn_dt_entry.pack(side="left", padx=(0, 8))

        self._txn_show_cancelled = tk.BooleanVar(value=False)
        ctk.CTkCheckBox(top, text="Show Cancelled",
                        variable=self._txn_show_cancelled).pack(side="left", padx=8)

        row2 = ctk.CTkFrame(parent, fg_color="transparent")
        row2.pack(fill="x", pady=(0, 6))

        ctk.CTkLabel(row2, text="Vendor:").pack(side="left", padx=(8, 4))
        vendors = get_vendors(self.db_path)
        self._txn_vnames = [v["name"] for v in vendors]
        self._txn_vids   = [v["id"]   for v in vendors]
        self._txn_vmenu = SearchableComboBox(row2, values=self._txn_vnames, width=220)
        self._txn_vmenu.pack(side="left", padx=4)

        ctk.CTkButton(row2, text="Show Report", width=110, fg_color="#1B4FD8",
                      command=self._load_transactions).pack(side="left", padx=8)
        ctk.CTkButton(row2, text="Export Excel", width=110, fg_color="#16A34A",
                      command=self._export_transactions).pack(side="left", padx=4)

        self._txn_summary = ctk.CTkLabel(parent, text="Choose a report type and click Show Report",
                                         font=ctk.CTkFont(size=12), text_color="#888")
        self._txn_summary.pack(anchor="w", padx=8, pady=(0, 6))

        self._txn_table_container = ctk.CTkFrame(parent, fg_color="transparent")
        self._txn_table_container.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        self._on_txn_type_change(self._txn_type_var.get())

    def _on_txn_type_change(self, _choice):
        cfg = self._txn_configs[self._txn_type_var.get()]
        state = "normal" if cfg["vendor_filter"] else "disabled"
        try:
            self._txn_vmenu.configure(state=state)
        except Exception:
            pass
        self._build_txn_table()

    def _build_txn_table(self):
        for w in self._txn_table_container.winfo_children():
            w.destroy()
        cfg = self._txn_configs[self._txn_type_var.get()]

        hdr = ctk.CTkFrame(self._txn_table_container, fg_color="#1B2B6B", corner_radius=6)
        hdr.pack(fill="x", pady=(0, 0))
        for col, w in zip(cfg["headers"], cfg["widths"]):
            ctk.CTkLabel(hdr, text=col, width=w,
                         font=ctk.CTkFont(size=12, weight="bold"),
                         text_color="white").pack(side="left", padx=4, pady=7)

        self._txn_scroll = ctk.CTkScrollableFrame(self._txn_table_container, fg_color="transparent")
        self._txn_scroll.pack(fill="both", expand=True, pady=(0, 4))

    def _load_transactions(self):
        cfg = self._txn_configs[self._txn_type_var.get()]
        self._build_txn_table()

        df = self._txn_df_entry.get() or None
        dt = self._txn_dt_entry.get() or None
        vid = None
        if cfg["vendor_filter"]:
            name = self._txn_vmenu.get()
            if name and name in self._txn_vnames:
                vid = self._txn_vids[self._txn_vnames.index(name)]

        rows = cfg["fetch"](df, dt, vid, self._txn_show_cancelled.get())
        self._txn_rows = rows
        self._txn_cfg_name = self._txn_type_var.get()

        active = [r for r in rows if r[-1] != "cancelled"]
        self._txn_summary.configure(
            text=f"{self._txn_type_var.get()}  |  {len(rows)} entries ({len(active)} active)",
            text_color="#1B2B6B")

        if not rows:
            ctk.CTkLabel(self._txn_scroll, text="No entries found.",
                         text_color="#999").pack(pady=20)
            return

        numeric_idx = cfg["numeric"]
        for i, vals in enumerate(rows):
            is_cancelled = vals[-1] == "cancelled"
            bg = "#FFF0F0" if is_cancelled else ("#F8FAFF" if i % 2 == 0 else "white")
            row = ctk.CTkFrame(self._txn_scroll, fg_color=bg, corner_radius=4)
            row.pack(fill="x", pady=1)
            color = "#888" if is_cancelled else "#333"
            for j, (val, width) in enumerate(zip(vals, cfg["widths"])):
                disp = fmt(val) if j in numeric_idx else str(val)[:40]
                ctk.CTkLabel(row, text=disp, width=width,
                             font=ctk.CTkFont(size=11), text_color=color
                             ).pack(side="left", padx=4, pady=5)

        foot = ctk.CTkFrame(self._txn_scroll, fg_color="#1B2B6B", corner_radius=4)
        foot.pack(fill="x", pady=(4, 2))
        foot_vals = []
        for j, (col, width) in enumerate(zip(cfg["headers"], cfg["widths"])):
            if j == 0:
                foot_vals.append("TOTAL")
            elif j in numeric_idx:
                foot_vals.append(fmt(sum(r[j] for r in rows)))
            else:
                foot_vals.append("")
        for val, width in zip(foot_vals, cfg["widths"]):
            ctk.CTkLabel(foot, text=str(val), width=width,
                         font=ctk.CTkFont(size=11, weight="bold"),
                         text_color="#7FFFD4").pack(side="left", padx=4, pady=6)

    def _export_transactions(self):
        if not self._txn_rows:
            messagebox.showwarning("No Data", "Load a report first.")
            return
        cfg = self._txn_configs.get(getattr(self, "_txn_cfg_name", self._txn_type_var.get()))
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel", "*.xlsx")],
                                            title=f"Save {self._txn_cfg_name}")
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self._txn_cfg_name[:31]
            ws.append([self._txn_cfg_name])
            ws.append(cfg["headers"])
            for vals in self._txn_rows:
                ws.append(vals)
            numeric_idx = cfg["numeric"]
            foot = ["" for _ in cfg["headers"]]
            foot[0] = "TOTAL"
            for j in numeric_idx:
                foot[j] = sum(r[j] for r in self._txn_rows)
            ws.append(foot)
            wb.save(path)
            messagebox.showinfo("Exported", f"{self._txn_cfg_name} saved:\n{path}")
        except Exception as ex:
            messagebox.showerror("Error", str(ex))

    # ── Outstanding Report ──────────────────────────────────────────────────

    def _build_outstanding(self, parent):
        top = ctk.CTkFrame(parent, fg_color="transparent")
        top.pack(fill="x", pady=(10, 6))

        ctk.CTkLabel(top, text="Vendor:").pack(side="left", padx=(8, 4))
        vendors = get_vendors(self.db_path, category="creditor")
        self._out_vnames = [v["name"] for v in vendors]
        self._out_vids   = [v["id"]   for v in vendors]
        self._out_vmenu = SearchableComboBox(top, values=self._out_vnames, width=240)
        self._out_vmenu.pack(side="left", padx=4)

        ctk.CTkButton(top, text="Show Report", width=110, fg_color="#1B4FD8",
                      command=self._load_outstanding).pack(side="left", padx=8)
        ctk.CTkButton(top, text="Export Excel", width=110, fg_color="#16A34A",
                      command=self._export_outstanding).pack(side="left", padx=4)

        self._out_summary = ctk.CTkLabel(parent,
                                          text="Select a vendor and click Show Report",
                                          font=ctk.CTkFont(size=12), text_color="#888")
        self._out_summary.pack(anchor="w", padx=8, pady=(0, 4))

        self._out_chips = ctk.CTkFrame(parent, fg_color="#EEF2FF", corner_radius=8)
        self._out_chips.pack(fill="x", padx=8, pady=(0, 6))
        self._out_chips_inner = ctk.CTkFrame(self._out_chips, fg_color="transparent")
        self._out_chips_inner.pack(fill="x", padx=12, pady=8)

        hdr = ctk.CTkFrame(parent, fg_color="#1B2B6B", corner_radius=6)
        hdr.pack(fill="x", padx=8, pady=(0, 0))
        out_hdrs = ["Date", "Invoice / Ref No", "Description",
                    "Debit (↓ Payable)", "Credit (↑ Payable)", "Balance"]
        out_ws   = [90, 140, 260, 140, 140, 120]
        self._out_widths = out_ws
        for col, w in zip(out_hdrs, out_ws):
            ctk.CTkLabel(hdr, text=col, width=w,
                         font=ctk.CTkFont(size=12, weight="bold"),
                         text_color="white").pack(side="left", padx=4, pady=7)

        self._out_scroll = ctk.CTkScrollableFrame(parent, fg_color="transparent")
        self._out_scroll.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        self._out_rows = []
        self._out_name = ""

    def _load_outstanding(self):
        name = self._out_vmenu.get()
        if not name or name not in self._out_vnames:
            messagebox.showerror("Error", "Please select a vendor.")
            return
        vid = self._out_vids[self._out_vnames.index(name)]
        data = get_vendor_outstanding_report(self.db_path, vid, self.fy_id)
        self._out_rows = data["rows"]
        self._out_name = data["name"]

        net = data["net_balance"]
        bal_tag = "PAYABLE" if net >= 0 else "ADVANCE"
        self._out_summary.configure(
            text=(f"{data['name']}  |  Total Debit: {fmt(data['total_debit'])}"
                  f"  |  Total Credit: {fmt(data['total_credit'])}"
                  f"  |  Net {bal_tag}: {fmt(abs(net))}"
                  f"  |  {len(self._out_rows)} invoice(s)"),
            text_color="#1B2B6B"
        )

        for w in self._out_chips_inner.winfo_children():
            w.destroy()
        chip_data = [
            ("Total Debit",   data["total_debit"],  "#D97706"),
            ("Total Credit",  data["total_credit"], "#1B2B6B"),
            ("Net Balance",
             f"{'PAYABLE' if net >= 0 else 'ADVANCE'}: {fmt(abs(net))}",
             "#EF4444" if net < 0 else "#1B4FD8"),
        ]
        for label, val, color in chip_data:
            frm = ctk.CTkFrame(self._out_chips_inner, fg_color="transparent")
            frm.pack(side="left", padx=20)
            ctk.CTkLabel(frm, text=label, font=ctk.CTkFont(size=11), text_color="#555").pack()
            ctk.CTkLabel(frm, text=fmt(val) if isinstance(val, (int, float)) else val,
                         font=ctk.CTkFont(size=14, weight="bold"), text_color=color).pack()

        for w in self._out_scroll.winfo_children():
            w.destroy()
        if not self._out_rows:
            ctk.CTkLabel(self._out_scroll, text="No invoices found.", text_color="#999").pack(pady=20)
            return
        for i, e in enumerate(self._out_rows):
            is_ob = e.get("type") == "Opening Balance"
            bg = "#EEF2FF" if is_ob else ("#F8FAFF" if i % 2 == 0 else "white")
            row = ctk.CTkFrame(self._out_scroll, fg_color=bg, corner_radius=4)
            row.pack(fill="x", pady=1)
            bal_color = "#EF4444" if e["balance"] < 0 else "#1B4FD8"
            date_str = to_display(e["date"]) if e.get("date") else "-"
            vals = [date_str,
                    str(e.get("ref") or "-")[:20],
                    str(e.get("description") or "-")[:38],
                    fmt(e["debit"]), fmt(e["credit"]), fmt(e["balance"])]
            for j, (val, width) in enumerate(zip(vals, self._out_widths)):
                vc = bal_color if j == 5 else ("#1B4FD8" if is_ob else "#333")
                ctk.CTkLabel(row, text=str(val), width=width,
                             font=ctk.CTkFont(size=11), text_color=vc).pack(side="left", padx=4, pady=5)

        total_cr = sum(r["credit"] for r in self._out_rows)
        total_dr = sum(r["debit"]  for r in self._out_rows)
        net_f = total_cr - total_dr
        foot = ctk.CTkFrame(self._out_scroll, fg_color="#1B2B6B", corner_radius=4)
        foot.pack(fill="x", pady=(4, 2))
        foot_vals  = ["", "TOTAL", "", fmt(total_dr), fmt(total_cr), fmt(net_f)]
        foot_clrs  = ["white", "white", "white", "#FFA07A", "#7FFFD4",
                      "#EF4444" if net_f < 0 else "#7FFFD4"]
        for val, width, color in zip(foot_vals, self._out_widths, foot_clrs):
            ctk.CTkLabel(foot, text=str(val), width=width,
                         font=ctk.CTkFont(size=11, weight="bold"),
                         text_color=color).pack(side="left", padx=4, pady=6)

    def _export_outstanding(self):
        if not self._out_rows:
            messagebox.showwarning("No Data", "Load the Outstanding Report first.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel", "*.xlsx")],
                                            title="Save Outstanding Report")
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Outstanding"
            ws.append([f"Outstanding Report: {self._out_name}"])
            ws.append(["Date", "Invoice / Ref No", "Description",
                       "Debit (↓ Payable)", "Credit (↑ Payable)", "Balance"])
            for e in self._out_rows:
                ws.append([to_display(e["date"]) if e.get("date") else "-",
                           e.get("ref") or "-", e.get("description") or "",
                           e["debit"], e["credit"], e["balance"]])
            td = sum(r["debit"]  for r in self._out_rows)
            tc = sum(r["credit"] for r in self._out_rows)
            ws.append(["", "TOTAL", "", td, tc, tc - td])
            wb.save(path)
            messagebox.showinfo("Exported", f"Outstanding Report saved:\n{path}")
        except Exception as ex:
            messagebox.showerror("Error", str(ex))

    # ── Invoice Settlement Report ────────────────────────────────────────────

    def _build_settlement(self, parent):
        top = ctk.CTkFrame(parent, fg_color="transparent")
        top.pack(fill="x", pady=(10, 6))

        ctk.CTkLabel(top, text="Vendor:").pack(side="left", padx=(8, 4))
        vendors = get_vendors(self.db_path, category="creditor")
        self._stl_vnames = [v["name"] for v in vendors]
        self._stl_vids   = [v["id"]   for v in vendors]
        self._stl_vmenu = SearchableComboBox(top, values=self._stl_vnames,
                                             command=self._load_invoices, width=220)
        self._stl_vmenu.pack(side="left", padx=4)

        ctk.CTkLabel(top, text="Invoice:").pack(side="left", padx=(10, 4))
        self._stl_inames = []
        self._stl_iids   = []
        self._stl_imenu = SearchableComboBox(top, values=self._stl_inames, width=200)
        self._stl_imenu.pack(side="left", padx=4)

        ctk.CTkButton(top, text="Show Settlement", width=120, fg_color="#1B4FD8",
                      command=self._load_settlement).pack(side="left", padx=8)
        ctk.CTkButton(top, text="Export Excel", width=110, fg_color="#16A34A",
                      command=self._export_settlement).pack(side="left", padx=4)

        self._stl_summary = ctk.CTkLabel(parent,
                                          text="Select a vendor, then an invoice, then click Show Settlement",
                                          font=ctk.CTkFont(size=12), text_color="#888")
        self._stl_summary.pack(anchor="w", padx=8, pady=(0, 4))

        hdr = ctk.CTkFrame(parent, fg_color="#1B2B6B", corner_radius=6)
        hdr.pack(fill="x", padx=8)
        stl_hdrs = ["Type", "Voucher / Ref", "Date", "Narration",
                    "Debit (↓ Payable)", "Credit (↑ Payable)", "Balance"]
        stl_ws   = [110, 120, 90, 240, 130, 130, 120]
        self._stl_widths = stl_ws
        for col, w in zip(stl_hdrs, stl_ws):
            ctk.CTkLabel(hdr, text=col, width=w,
                         font=ctk.CTkFont(size=12, weight="bold"),
                         text_color="white").pack(side="left", padx=4, pady=7)

        self._stl_scroll = ctk.CTkScrollableFrame(parent, fg_color="transparent")
        self._stl_scroll.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        self._stl_rows = []
        self._stl_inv_name = ""

    def _load_invoices(self, *_):
        name = self._stl_vmenu.get()
        if not name or name not in self._stl_vnames:
            return
        vid = self._stl_vids[self._stl_vnames.index(name)]
        purchases = get_purchases(self.db_path, vendor_id=vid)
        self._stl_inames = [
            p.get("invoice_number") or p["voucher_no"] for p in purchases
        ]
        self._stl_iids = [p["id"] for p in purchases]
        self._stl_imenu.configure(values=self._stl_inames)
        self._stl_imenu.set("")

    def _load_settlement(self):
        vname = self._stl_vmenu.get()
        iname = self._stl_imenu.get()
        if not vname or vname not in self._stl_vnames:
            messagebox.showerror("Error", "Please select a vendor.")
            return
        if not iname or iname not in self._stl_inames:
            messagebox.showerror("Error", "Please select an invoice.")
            return

        vid = self._stl_vids[self._stl_vnames.index(vname)]
        pid = self._stl_iids[self._stl_inames.index(iname)]

        rows = []
        balance = 0

        purchases = get_purchases(self.db_path, vendor_id=vid)
        purch = next((p for p in purchases if p["id"] == pid), None)
        if not purch:
            messagebox.showerror("Error", "Invoice not found.")
            return
        self._stl_inv_name = iname
        inv_ref = purch.get("invoice_number") or purch["voucher_no"]
        pv = purch["total_value"]
        balance += pv
        rows.append({"type": "Purchase", "ref": purch["voucher_no"],
                     "date": purch["date"],
                     "narration": purch.get("narration") or f"Purchase | Inv: {inv_ref}",
                     "debit": 0, "credit": pv, "balance": balance})

        cn_list = get_notes(self.db_path, "credit_notes")
        for cn in cn_list:
            if cn.get("ref_purchase_id") == pid:
                balance += cn["total_value"]
                rows.append({"type": "Credit Note", "ref": cn["voucher_no"],
                             "date": cn["date"],
                             "narration": cn.get("narration") or "Credit Note",
                             "debit": 0, "credit": cn["total_value"], "balance": balance})

        dn_list = get_notes(self.db_path, "debit_notes")
        for dn in dn_list:
            if dn.get("ref_purchase_id") == pid:
                balance -= dn["total_value"]
                rows.append({"type": "Debit Note", "ref": dn["voucher_no"],
                             "date": dn["date"],
                             "narration": dn.get("narration") or "Debit Note",
                             "debit": dn["total_value"], "credit": 0, "balance": balance})

        from database import get_connection
        conn = get_connection(self.db_path)
        cur = conn.cursor()
        cur.execute("""
            SELECT pa.amount, v.voucher_no, v.date, v.narration
            FROM payment_adjustments pa
            JOIN vouchers v ON pa.voucher_id=v.id
            WHERE pa.purchase_id=? AND v.status='active'
        """, (pid,))
        for pay in [dict(r) for r in cur.fetchall()]:
            balance -= pay["amount"]
            rows.append({"type": "Payment", "ref": pay["voucher_no"],
                         "date": pay["date"],
                         "narration": pay.get("narration") or "Payment",
                         "debit": pay["amount"], "credit": 0, "balance": balance})
        conn.close()

        self._stl_rows = rows

        total_cr = sum(r["credit"] for r in rows)
        total_dr = sum(r["debit"]  for r in rows)
        net = total_cr - total_dr
        bal_tag = "PAYABLE" if net >= 0 else "SETTLED / ADVANCE"
        self._stl_summary.configure(
            text=(f"Invoice: {inv_ref}  |  {vname}  |  "
                  f"Total Debit: {fmt(total_dr)}  |  Total Credit: {fmt(total_cr)}  |  {bal_tag}: {fmt(abs(net))}"),
            text_color="#1B2B6B"
        )

        for w in self._stl_scroll.winfo_children():
            w.destroy()
        for i, e in enumerate(rows):
            bg = "#F8FAFF" if i % 2 == 0 else "white"
            row_frame = ctk.CTkFrame(self._stl_scroll, fg_color=bg, corner_radius=4)
            row_frame.pack(fill="x", pady=1)
            bal_color = "#EF4444" if e["balance"] < 0 else "#1B4FD8"
            type_colors = {"Purchase": "#1B2B6B", "Credit Note": "#7C3AED",
                           "Debit Note": "#D97706", "Payment": "#16A34A"}
            type_color = type_colors.get(e["type"], "#333")
            date_str = to_display(e["date"]) if e.get("date") else "-"
            vals = [e["type"], e["ref"], date_str,
                    str(e["narration"])[:35],
                    fmt(e["debit"]), fmt(e["credit"]), fmt(e["balance"])]
            for j, (val, width) in enumerate(zip(vals, self._stl_widths)):
                vc = type_color if j == 0 else (bal_color if j == 6 else "#333")
                ctk.CTkLabel(row_frame, text=str(val), width=width,
                             font=ctk.CTkFont(size=11), text_color=vc).pack(side="left", padx=4, pady=5)

        tc = sum(r["credit"] for r in rows)
        td = sum(r["debit"]  for r in rows)
        nf = tc - td
        foot = ctk.CTkFrame(self._stl_scroll, fg_color="#1B2B6B", corner_radius=4)
        foot.pack(fill="x", pady=(4, 2))
        foot_vals = ["TOTAL", "", "", "", fmt(td), fmt(tc), fmt(nf)]
        foot_clrs = ["white", "white", "white", "white", "#FFA07A", "#7FFFD4",
                     "#EF4444" if nf < 0 else "#7FFFD4"]
        for val, width, color in zip(foot_vals, self._stl_widths, foot_clrs):
            ctk.CTkLabel(foot, text=str(val), width=width,
                         font=ctk.CTkFont(size=11, weight="bold"),
                         text_color=color).pack(side="left", padx=4, pady=6)

    def _export_settlement(self):
        if not self._stl_rows:
            messagebox.showwarning("No Data", "Load the Settlement Report first.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel", "*.xlsx")],
                                            title="Save Settlement Report")
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Settlement"
            ws.append([f"Invoice Settlement: {self._stl_inv_name}"])
            ws.append(["Type", "Voucher / Ref", "Date", "Narration",
                       "Debit (↓ Payable)", "Credit (↑ Payable)", "Balance"])
            for e in self._stl_rows:
                ws.append([e["type"], e["ref"],
                           to_display(e["date"]) if e.get("date") else "-",
                           e["narration"], e["debit"], e["credit"], e["balance"]])
            td = sum(r["debit"]  for r in self._stl_rows)
            tc = sum(r["credit"] for r in self._stl_rows)
            ws.append(["TOTAL", "", "", "", td, tc, tc - td])
            wb.save(path)
            messagebox.showinfo("Exported", f"Settlement Report saved:\n{path}")
        except Exception as ex:
            messagebox.showerror("Error", str(ex))

    def refresh(self):
        txn_vendors = get_vendors(self.db_path)
        self._txn_vnames = [v["name"] for v in txn_vendors]
        self._txn_vids   = [v["id"]   for v in txn_vendors]
        self._txn_vmenu.configure(values=self._txn_vnames)
        self._txn_vmenu.set("")

        vendors = get_vendors(self.db_path, category="creditor")
        vnames = [v["name"] for v in vendors]
        vids   = [v["id"]   for v in vendors]
        self._out_vnames = vnames
        self._out_vids   = vids
        self._out_vmenu.configure(values=vnames)
        self._out_vmenu.set("")

        self._stl_vnames = vnames
        self._stl_vids   = vids
        self._stl_vmenu.configure(values=vnames)
        self._stl_vmenu.set("")
        self._stl_inames = []
        self._stl_iids   = []
        self._stl_imenu.configure(values=[])
        self._stl_imenu.set("")
